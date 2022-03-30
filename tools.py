
import os
import time
import uuid
import hashlib
import json
import shutil


import openpyxl
from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import List, Optional
from fastapi import Depends, APIRouter, File, Form, Path, Request, Query
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def get_timestamp():
    return round(time.time()*1000)

def create_dir(dir_path):
    if not os.path.isdir(dir_path):
        os.makedirs(dir_path)
        return 200,""
    else:
        return 400,"创建失败，因为目标文件夹已存在"


def get_hashed_text(text: str = None, mix_text: str = None):
    text = str(text)
    if not bool(text):
        text = "{}{}".format(str(get_timestamp()), uuid.uuid1().hex)
    if bool(mix_text):
        text = "{}{}".format(str(text), str(mix_text))
    return hashlib.sha256(text.encode()).hexdigest()


def get_hashed_file(file: bytes = None):
    return hashlib.sha256(file).hexdigest()


def get_uuid_text():
    return uuid.uuid1().hex


# 准备相关目录
work_path = os.path.dirname(__file__)
print("boot dir is: ",work_path)
cache_path = os.path.join(work_path,"cache")
users_path = os.path.join(cache_path,"users")
sheets_path = os.path.join(cache_path,"sheets")
required_paths = [cache_path,users_path,sheets_path]
for path in required_paths:
    create_dir(path)


def get_user_path(user_id):
    user_path = os.path.join(users_path,user_id)
    if not os.path.isdir(user_path):
        os.makedirs(user_path)
    return user_path


def get_sheet_path(sheet_id):
    return os.path.join(sheets_path,sheet_id)


def init_sheet_path(sheet_id):
    path = get_sheet_path(sheet_id)
    if not os.path.isdir(path):
        os.makedirs(path)
    return path

def get_db_path(sheet_id):
    return os.path.join(get_sheet_path(sheet_id), f"{sheet_id}.db")
    

def init_sheet_path_and_id():
    sheet_id = get_uuid_text()
    sheet_path = init_sheet_path(sheet_id)
    return sheet_id,sheet_path


def delete_sheet_path_and_files(sheet_id):
    sheet_path = get_sheet_path(sheet_id)
    shutil.rmtree(sheet_path)


def get_matched_data(data:dict={},template:dict={}):
    matched_key = data.keys()&template.keys()
    bad_keys = set(data.keys()).difference(set(matched_key))
    for k in bad_keys:
        del data[k]
    return data

def parse_sheet_columns_from_xlsx_file(xlsx_file: bytes, sheet_id:str):
    """
    return 200,[{"type": str,"name": str,"link_to": str}]
    """
    try:
        # 读取 xlsx 文件
        workbook = load_workbook(BytesIO(xlsx_file))
        # 取出全部数据
        data = {}
        for row in workbook.worksheets[0].rows:
            for cell in row:
                data[cell.coordinate] = {"value": cell.value}
        # 遍历，取出数据标签
        columns = {}
        for cell in data:
            value = data[cell].get("value")
            if not value:
                continue
            # 当遇到形如 input:str->A2 的格子时，标记为需要采集的格子
            mark_a = "->"
            mark_b = "input:"
            if mark_a in value and mark_b in value:
                type, link_to = value[len(mark_b):].split(mark_a)
                columns[cell] = {
                    "type": type,
                    "name": data[link_to].get("value"),
                    "link_to": link_to,
                }
        if not columns:
            return 400,"未匹配到表单标识"
        # 生成缓存
        with open(os.path.join(get_sheet_path(sheet_id), "columns.json"), "wt", encoding="utf-8")as t:
            t.write(json.dumps(columns))
        return 200,columns
    except Exception as e:
        return 500,str(e)


def get_sheet_columns_from_cache(sheet_id) -> dict:
    """
    数据列是 "空格key":{指向格信息}
    """
    with open(os.path.join(get_sheet_path(sheet_id), "columns.json"), "rt", encoding="utf-8")as t:
        return json.loads(t.read())


def get_sheet_data_template(sheet_id) -> dict:
    # 数据模型是与数据库一致的模型,用 指向位置 做键
    template = {}
    columns = get_sheet_columns_from_cache(sheet_id)
    for key in columns.keys():
        template[columns[key]["link_to"]] = {"input_from":key,"type":columns[key]["type"],"name":columns[key]["name"]}
    return template


def add_sheet_to_user_status(user_id,sheet:list={"index": str, "title": str}):
    user_path = get_user_path(user_id)
    user_json = os.path.join(user_path, "status.json")
    data = {"sheets":{}}
    if os.path.isfile(user_json):
        with open(user_json, "rt", encoding="utf-8")as t:
            data = json.loads(t.read())
    
    data["sheets"][sheet["index"]] =  {"title": sheet["title"], "create_at": get_timestamp()}
    with open(user_json, "wt", encoding="utf-8")as t:
        t.write(json.dumps(data))



def delete_sheet_from_user_status(user_id,sheet_id):
    user_path = get_user_path(user_id)
    user_json = os.path.join(user_path, "status.json")
    
    if os.path.isfile(user_json):
        with open(user_json, "rt", encoding="utf-8")as t:
            data = json.loads(t.read())
            del data["sheets"][sheet_id]
            with open(user_json, "wt", encoding="utf-8")as t2:
                t2.write(json.dumps(data))



def get_sheets_from_user_status(user_id):
    user_path = get_user_path(user_id)
    user_json = os.path.join(user_path, "status.json")

    if os.path.isfile(user_json):
        with open(user_json, "rt", encoding="utf-8")as t:
            data = json.loads(t.read())
            return 200,data["sheets"]
    else:
        return 200,[]
    
    



def add_subusers_to_sheet(user_id, sheet_id, users):
    # 检查 表单所属
    user_path = get_user_path(user_id)
    code,result = get_sheets_from_user_status(user_id)
    if code !=200:
        return code,result
    if sheet_id not in result:
        return 401,"只能向拥有的表单导入子用户"
    # 写入信息
    sheet_path = get_sheet_path(sheet_id)
    sheet_subusers_path = os.path.join(sheet_path, "subusers.json")
    with open(sheet_subusers_path, "wt", encoding="utf-8")as t:
        t.write(json.dumps(users))
    return 200,"导入成功"




def check_sheet_subuser(subuser_id, password_hash, sheet_id):

    sheet_path = get_sheet_path(sheet_id)
    subusers_path = os.path.join(sheet_path, "subusers.json")
    if not os.path.isfile(subusers_path):
        return False
    with open(subusers_path, "rt", encoding="utf-8")as t:
        users: dict = json.loads(t.read())
        user: dict = users.get(subuser_id)
        if not user:
            return False
        if user.get("p") == password_hash:
            return True
        else:
            return False




# def get_db_columns(form_id):
#     marked_data = get_sheet_columns_from_cache(form_id)
#     db_columns = [
#         {
#             "type": marked_data[i]["type"],
#             "column_name":marked_data[i]["link_to"],
#             "defalt_value":marked_data[i]["defalt_value"]
#         } for i in marked_data
#     ]
#     return db_columns


# def get_db_columns_name(form_id):
#     return [i["column_name"] for i in get_db_columns(form_id)]


# def get_map_of_link_to(form_id):
#     r = {}
#     columns = get_sheet_columns_from_cache(form_id)
#     for i in columns.keys():
#         r[columns[i]["link_to"]] = i
#     return r


def fill_to_xlsx(sheet_id, row: dict):
    print("output:", sheet_id, row)
    try:
        workbook = load_workbook(os.path.join(
            get_sheet_path(sheet_id), "{}.xlsx".format(sheet_id)))
        sheet = workbook.worksheets[0]
        template = get_sheet_data_template(sheet_id)
        data = get_matched_data(data=row,template=template)
        for key in data.keys():
            sheet[template[key]["input_from"]] = row[key]
        out = BytesIO()
        workbook.save(out)
        out.seek(0)
        return 200, out
    except Exception as e:
        return 500, str(e)


# def append_form_to_user(user_id, form_id, form_title):
#     user_path = get_user_path(user_id)
#     forms_path = os.path.join(user_path, "forms.json")
#     temp_forms = []
#     if os.path.isfile(forms_path):
#         with open(forms_path, "rt", encoding="utf-8")as t:
#             temp_forms += json.loads(t.read())
#     temp_forms.append(
#         {"form_id": form_id, "form_title": form_title, "create_at": time.time()})
#     with open(forms_path, "wt", encoding="utf-8")as t:
#         t.write(json.dumps(temp_forms))

