

import os
import json
import time
from io import StringIO, BytesIO
from tempfile import NamedTemporaryFile
from typing import Optional
from fastapi import Depends, APIRouter, File, Form, Path, Request, Query
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


from tools import get_db_columns, get_db_columns_name,get_marked_data,fill_to_xlsx,check_form_user,append_form_to_user,append_users_to_form,init_form_path, cache_path, get_form_path, get_db_path, get_user_path, get_hashed_text,get_marked_data,get_hashed_file,get_map_of_link_to,get_marks_from_xlsx_sheet,get_uuid_text
from db.database_sql import create as create_db
from db.database_sql import insert as insert_db
from db.database_sql import update as update_db
from db.database_sql import delete as delete_db
from db.database_sql import rows as get_rows_from_db
from db.database_sql import row_by_id
from middlewares import get_user_by_api_token,get_user_by_api_user_token


router = APIRouter(prefix="/xlsx")




@router.post("/")
def create_form_by_xlsx(
    file: bytes = File(...),
    title: str = Form(...),
    user_info: list = Depends(get_user_by_api_token)

):
    if not file:
        return {"code": 400, "errmsg": "文件缺失"}

    user_id, _ = user_info
    # 初始化一个表单 ID 和文件夹
    form_id, form_path = init_form_path()
    # 添加表单到用户目录
    append_form_to_user(user_id, form_id, title)
    # 保存 xlsx 文件
    with open(os.path.join(form_path, f"{form_id}.xlsx"), "wb")as b:
        b.write(file)
    # 读取 xlsx 文件，并获取被标记的单元格
    workbook = load_workbook(BytesIO(file))
    marked_data = get_marks_from_xlsx_sheet(workbook.worksheets[0])
    if not marked_data:
        return {"code": 400, "errmsg": "未匹配到表单标识"}
    with open(os.path.join(form_path, f"{form_id}.json"), "wt", encoding="utf-8")as t:
        t.write(json.dumps(marked_data))
    # 创建数据库
    code, result = create_db(
        db_path=os.path.join(form_path, f"{form_id}.db"),
        db_columns=get_db_columns(form_id)
    )
    return {"code": 200, "db": [code, result], "form_id": form_id, "form_title": title, "marked_data": marked_data}


@router.post("/{form_id}/sheet/row")
async def insert_data(
    request: Request,
    user_info: list = Depends(get_user_by_api_user_token)
):
    form_id,user_id,password_hash = user_info
    
    form_path = get_form_path(form_id)
    if not os.path.isfile(os.path.join(form_path, f"{form_id}.db")):
        return {"code": 404, "errmsg": "表单数据库不存在"}

    data: dict = await request.json()

    db_columns_name = get_db_columns_name(form_id)
    print(db_columns_name)

    # 检查需要插入的数据字段是否在数据库中
    temp = []
    for name in data.keys():
        if name not in db_columns_name:
            temp.append(name)
    if temp:
        return {"code": 400, "errmsg": "illegal column: {} ".format(temp)}
    # 插入数据
    code, result = insert_db(
        db_path=get_db_path(form_id),
        row=data
    )
    return {"code": 200, "db": [code, result]}


@router.get("/{form_id}/sheet/rows")
async def get_form_rows(
    request: Request,
    limit: int = Query(default=10),
    offset: int = Query(default=0),
    user_info: list = Depends(get_user_by_api_user_token)
):
    form_id,user_id,password_hash = user_info
    form_path = get_form_path(form_id)
    if not os.path.isfile(os.path.join(form_path, f"{form_id}.db")):
        return {"code": 404, "errmsg": "表单数据库不存在"}

    code, result = get_rows_from_db(
        db_path=get_db_path(form_id),
        limit=limit,
        offset=offset
    )
    return {"code": 200, "db": [code, result]}


@router.get("/{form_id}/sheet/columns")
async def get_form_columns(
    user_info: list = Depends(get_user_by_api_user_token)
):
    form_id,user_id,password_hash = user_info
    form_path = get_form_path(form_id)
    if not os.path.isfile(os.path.join(form_path, f"{form_id}.json")):
        return {"code": 404, "errmsg": "表单结构数据不存在"}
    return {"code": 200, "data": get_db_columns(form_id)}


@router.post("/{form_id}/sheet/rows/{row_id}")
async def update_form_row(
    request: Request,
    row_id: int = Path(...),
    user_info: list = Depends(get_user_by_api_user_token)
):
    form_id,user_id,password_hash = user_info
    form_path = get_form_path(form_id)
    if not os.path.isfile(os.path.join(form_path, f"{form_id}.db")):
        return {"code": 404, "errmsg": "表单数据库不存在"}

    data: dict = await request.json()
    db_columns_name = get_db_columns_name(form_id)
    # 检查需要插入的数据字段是否在数据库中
    temp = []
    for name in data.keys():
        if name not in db_columns_name:
            temp.append(name)
    if temp:
        return {"code": 400, "errmsg": "illegal column: {} ".format(temp)}

    code, result = update_db(
        db_path=get_db_path(form_id),
        row_id=row_id,
        row=data
    )
    return {"code": 200, "db": [code, result]}


@router.get("/{form_id}/sheet/row/{row_id}")
async def get_form_row(
    row_id: int = Path(...),
    user_info: list = Depends(get_user_by_api_user_token)
):
    form_id,user_id,password_hash = user_info
    form_path = get_form_path(form_id)
    if not os.path.isfile(os.path.join(form_path, f"{form_id}.db")):
        return {"code": 404, "errmsg": "表单数据库不存在"}

    code, result = row_by_id(
        db_path=get_db_path(form_id),
        row_id=row_id
    )
    return {"code": 200, "db": [code, result]}



@router.delete("/{form_id}/sheet/rows/{row_id}")
async def delete_form_row(
    request: Request,
    row_id: int = Path(...),
    user_info: list = Depends(get_user_by_api_user_token)
):
    form_id,user_id,password_hash = user_info
    form_path = get_form_path(form_id)
    if not os.path.isfile(os.path.join(form_path, f"{form_id}.db")):
        return {"code": 404, "errmsg": "表单数据库不存在"}

    code, result = delete_db(
        db_path=get_db_path(form_id),
        row_id=row_id
    )
    return {"code": 200, "db": [code, result]}



@router.get("/{form_id}/sheet/rows/{row_id}/xlsx")
def get_form_row_to_xlsx(
    row_id: int = Path(...),
    user_info: list = Depends(get_user_by_api_user_token)
):
    form_id,user_id,password_hash = user_info
    form_path = get_form_path(form_id)
    if not os.path.isfile(os.path.join(form_path, f"{form_id}.db")):
        return {"code": 404, "errmsg": "表单数据库不存在"}

    code, result = row_by_id(
        db_path=get_db_path(form_id),
        row_id=row_id
    )
    if result:
        code, result = fill_to_xlsx(form_id, result[0])
        # with open(os.path.join(form_path,"{}.xlsx".format(row_id)),"wb")as t:
        #     t.write(result.read())
        if code == 200:
            headers = {
                'Content-Disposition': 'attachment; filename="{}-{}.xlsx"'.format(form_id, row_id)
            }
            return StreamingResponse(result,
                                     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                     headers=headers
                                     )

    return {"code": 200, "db": [code, result]}




@router.post("/{form_id}/users")
def create_form_users_by_xlsx(
    file: bytes = File(...),
    form_id: str = Path(...),
    user_info: list = Depends(get_user_by_api_token)
):
    if not file:
        return {"code": 400, "errmsg": "文件缺失"}

    user_id, _ = user_info

    users = {}

    workbook = load_workbook(BytesIO(file))
    worksheet = workbook.worksheets[0]
    for index, row in enumerate(worksheet.rows):
        if index == 0:
            continue
        users[get_hashed_text(row[0].value)] = {
            "a": row[0].value, "p": get_hashed_text(row[1].value)}

    append_users_to_form(user_id, form_id, users)
    return {"code": 200, "total": len(users)}


@router.post("/{form_id}/user/login")
async def form_users_login(
    request:Request,
    # account: str = Form(...),
    # password: str = Form(...),
    form_id: str = Path(...),
):
    data: dict = await request.json()
    print(data)
    account = data.get("account")
    password = data.get("password")
    user_id = get_hashed_text(account)
    password_hash = get_hashed_text(password)
    if check_form_user(user_id, password_hash, form_id):
        return {"code": 200, "x-api-user-code": "{}-{}".format(user_id, password_hash),"name":account}
    else:
        return {"code": 401, "errmsg": "校验失败"}
